import io
from typing import Any, List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from openpyxl import load_workbook
from app.core import security
from app.models.users import User
from app.schemas.user import UserOut, BulkUserUploadResponse
from app.api.deps import get_current_admin_user

router = APIRouter()

VALID_ROLES = {"admin", "faculty", "hod", "student", "deo"}
EXPECTED_COLUMNS = ["full_name", "username", "email", "password", "role"]


def validate_excel_headers(headers: list) -> bool:
    normalized = [str(h).strip().lower() for h in headers if h is not None]
    return normalized == EXPECTED_COLUMNS


def parse_excel_file(file_bytes: bytes) -> List[dict]:
    workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty.")

    headers = rows[0]
    if not validate_excel_headers(list(headers)):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid columns. Expected: {EXPECTED_COLUMNS}. "
                   f"Got: {[str(h).strip().lower() for h in headers if h]}",
        )

    records = []
    for row_num, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record = {}
        for col_idx, col_name in enumerate(EXPECTED_COLUMNS):
            value = row[col_idx] if col_idx < len(row) else None
            record[col_name] = str(value).strip() if value is not None else ""
        record["_row"] = row_num
        records.append(record)

    workbook.close()
    return records


def validate_record(record: dict) -> List[str]:
    errors = []
    if not record.get("full_name"):
        errors.append("full_name is required")
    if not record.get("username"):
        errors.append("username is required")
    if not record.get("email") or "@" not in record.get("email", ""):
        errors.append("a valid email is required")
    if not record.get("password") or len(record.get("password", "")) < 6:
        errors.append("password must be at least 6 characters")
    if record.get("role", "").lower() not in VALID_ROLES:
        errors.append(f"role must be one of {VALID_ROLES}, got '{record.get('role')}'")
    return errors


@router.post("/create-users", response_model=BulkUserUploadResponse)
async def create_users(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_admin_user),
):
    """
    Admin uploads an Excel (.xlsx) file with columns:
    full_name, username, email, password, role
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted.")

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Max 5 MB.")

    records = parse_excel_file(contents)
    if not records:
        raise HTTPException(status_code=400, detail="No data rows found in the file.")

    # Pre-fetch existing emails to detect duplicates
    existing_users = await User.find_all().to_list()
    existing_emails = {u.email.lower() for u in existing_users}

    seen_emails = set()
    successfully_created = 0
    created_users = []
    error_list = []

    for record in records:
        row_num = record.pop("_row")

        validation_errors = validate_record(record)
        if validation_errors:
            error_list.append({"row": row_num, "errors": validation_errors})
            continue

        email_lower = record["email"].lower()

        if email_lower in existing_emails:
            error_list.append({"row": row_num, "errors": [f"email '{record['email']}' already exists in database"]})
            continue

        if email_lower in seen_emails:
            error_list.append({"row": row_num, "errors": [f"duplicate email '{record['email']}' in file"]})
            continue

        try:
            hashed_password = security.get_password_hash(record["password"])
            new_user = User(
                full_name=record["full_name"],
                email=record["email"],
                hashed_password=hashed_password,
                role=record["role"].lower(),
            )
            await new_user.insert()

            seen_emails.add(email_lower)
            existing_emails.add(email_lower)
            successfully_created += 1
            created_users.append(UserOut(
                id=str(new_user.id),
                email=new_user.email,
                full_name=new_user.full_name,
                role=new_user.role,
                is_active=new_user.is_active,
            ))
        except Exception as e:
            error_list.append({"row": row_num, "errors": [str(e)]})

    return BulkUserUploadResponse(
        total_records=len(records),
        successfully_created=successfully_created,
        failed=len(error_list),
        errors=error_list,
        created_users=created_users,
    )


@router.get("/users", response_model=List[UserOut])
async def list_all_users(
    current_user: User = Depends(get_current_admin_user),
):
    """List all users (admin only)."""
    users = await User.find_all().to_list()
    return [
        UserOut(
            id=str(u.id),
            email=u.email,
            full_name=u.full_name,
            role=u.role,
            is_active=u.is_active,
        )
        for u in users
    ]


@router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    current_user: User = Depends(get_current_admin_user),
):
    """Delete a user by ID (admin only)."""
    from beanie import PydanticObjectId
    user = await User.get(PydanticObjectId(user_id))
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if str(user.id) == str(current_user.id):
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    await user.delete()
    return {"message": "User deleted successfully"}